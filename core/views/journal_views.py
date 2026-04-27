"""
CISIS — Views для журнала образцов.

Содержит:
- journal_samples: основная страница журнала
- export_journal_xlsx: экспорт в XLSX
- journal_filter_options: AJAX каскадные фильтры
- save_column_preferences: AJAX сохранение столбцов
- save_sample_column_widths: AJAX сохранение ширин столбцов
- save_filter_preferences: AJAX сохранение последних применённых фильтров (v3.81.0)
- Вспомогательные функции фильтрации и отображения
"""

import json
from datetime import date, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.utils.timezone import localtime
from django.db import models
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.urls import reverse

from core.models import (
    Sample, Laboratory, WorkshopStatus, JournalColumn,
)
from .constants import (
    WORKSHOP_ROLES, QMS_ROLES,
    JOURNAL_DISPLAYABLE_COLUMNS, DISPLAYABLE_COLUMNS_DICT,
    DEFAULT_COLUMNS_BY_ROLE, FILTERABLE_COLUMNS, ITEMS_PER_PAGE,
)
from core.permissions import PermissionChecker, CAN_SEE_PENDING_VERIFICATION

# ⭐ v3.89.0: Роли с доступом к табу «Черновики» в журнале образцов.
# Совпадает с теми, кто видит кнопку «➕ Добавить» (т.е. может
# создавать образцы). Тестировщики и цех черновиков не видят.
DRAFTS_VISIBLE_ROLES = ('CLIENT_MANAGER', 'CLIENT_DEPT_HEAD', 'LAB_HEAD', 'SYSADMIN')

# ─────────────────────────────────────────────────────────────
# Вспомогательные функции
# ─────────────────────────────────────────────────────────────

def _get_user_visible_columns(user):
    """
    Возвращает список кодов столбцов, которые пользователь МОЖЕТ видеть
    (на основе role_permissions).
    """
    available = []
    for code, name in JOURNAL_DISPLAYABLE_COLUMNS:
        if PermissionChecker.can_view(user, 'SAMPLES', code):
            available.append(code)
    return available


def _get_user_selected_columns(user):
    """
    Возвращает список кодов столбцов, выбранных пользователем.
    Если нет сохранённых настроек — возвращает дефолт для роли.
    """
    prefs = user.ui_preferences or {}
    saved = prefs.get('journal_columns', {}).get('SAMPLES')

    if saved:
        available = set(_get_user_visible_columns(user))
        return [c for c in saved if c in available]

    role = user.role
    default = DEFAULT_COLUMNS_BY_ROLE.get(role, DEFAULT_COLUMNS_BY_ROLE['_default'])
    available = set(_get_user_visible_columns(user))
    return [c for c in default if c in available]


def _build_base_queryset(user):
    """
    Строит базовый queryset образцов.

    v3.17.0: видимость лабораторий через role_laboratory_access.
    - role_laboratory_access с lab=NULL → все лаборатории
    - role_laboratory_access с конкретными lab → эти лаборатории
    - Нет записей → fallback: user.laboratory + additional_laboratories
    - Workshop-роли: отдельная логика (manufacturing/workshop_status)
    """
    user_role = user.role

    # ── Workshop: фильтрация по manufacturing/workshop_status, не по лаборатории ──
    # ⭐ v3.89.0: status__in вместо одиночного exclude — добавлен DRAFT.
    if user_role == 'WORKSHOP_HEAD':
        return Sample.objects.filter(
            manufacturing=True
        ).exclude(status__in=['PENDING_VERIFICATION', 'DRAFT'])

    if user_role == 'WORKSHOP':
        return Sample.objects.filter(
            workshop_status__isnull=False
        ).exclude(status__in=['PENDING_VERIFICATION', 'DRAFT'])

    # ── Все остальные роли: доступ через role_laboratory_access ──
    visible_lab_ids = PermissionChecker.get_visible_laboratory_ids(user, 'SAMPLES')

    if visible_lab_ids is None:
        # None = все лаборатории (без фильтра)
        samples = Sample.objects.all()
    elif not visible_lab_ids:
        # Пустой набор = нет доступа
        return Sample.objects.none()
    else:
        # Конкретные лаборатории
        samples = Sample.objects.filter(laboratory_id__in=visible_lab_ids)

    # ── Фильтр статуса PENDING_VERIFICATION ──
    if user_role not in CAN_SEE_PENDING_VERIFICATION:
        samples = samples.exclude(status='PENDING_VERIFICATION')

    # ⭐ v3.89.0: Черновики никогда не показываются в основном журнале —
    # они доступны только через таб «📋 Черновики» (см. journal_samples).
    samples = samples.exclude(status='DRAFT')

    return samples

def _apply_filters(queryset, params, user):
    """
    Применяет фильтры из GET-параметров к queryset.
    Возвращает отфильтрованный queryset.
    """
    # --- Select фильтры (множественный выбор) ---
    status_values = params.getlist('status')
    if status_values:
        queryset = queryset.filter(status__in=status_values)

    workshop_status_values = params.getlist('workshop_status')
    if workshop_status_values:
        queryset = queryset.filter(workshop_status__in=workshop_status_values)

    laboratory_values = params.getlist('laboratory')
    if laboratory_values:
        queryset = queryset.filter(laboratory_id__in=laboratory_values)

    client_values = params.getlist('client')
    if client_values:
        queryset = queryset.filter(client_id__in=client_values)

    contract_values = params.getlist('contract')
    if contract_values:
        queryset = queryset.filter(contract_id__in=contract_values)

    standard_values = params.getlist('standards')
    if standard_values:
        queryset = queryset.filter(standards__id__in=standard_values)

    accreditation_values = params.getlist('accreditation_area')
    if accreditation_values:
        queryset = queryset.filter(accreditation_area_id__in=accreditation_values)

    test_type_values = params.getlist('test_type')
    if test_type_values:
        queryset = queryset.filter(test_type__in=test_type_values)

    # ⭐ v3.32.0: report_type — запятая-разделённый, ищем по contains
    report_type_values = params.getlist('report_type')
    if report_type_values:
        rt_q = Q()
        for rt in report_type_values:
            rt_q |= Q(report_type__contains=rt)
        queryset = queryset.filter(rt_q)

    further_movement_values = params.getlist('further_movement')
    if further_movement_values:
        queryset = queryset.filter(further_movement__in=further_movement_values)

    registered_by_values = params.getlist('registered_by')
    if registered_by_values:
        queryset = queryset.filter(registered_by_id__in=registered_by_values)

    verified_by_values = params.getlist('verified_by')
    if verified_by_values:
        if '__none__' in verified_by_values:
            other_ids = [v for v in verified_by_values if v != '__none__']
            if other_ids:
                queryset = queryset.filter(
                    models.Q(verified_by__isnull=True) | models.Q(verified_by_id__in=other_ids)
                )
            else:
                queryset = queryset.filter(verified_by__isnull=True)
        else:
            queryset = queryset.filter(verified_by_id__in=verified_by_values)

    # --- Boolean фильтры ---
    manufacturing_val = params.get('manufacturing')
    if manufacturing_val in ('true', 'false'):
        queryset = queryset.filter(manufacturing=(manufacturing_val == 'true'))

    uzk_val = params.get('uzk_required')
    if uzk_val in ('true', 'false'):
        queryset = queryset.filter(uzk_required=(uzk_val == 'true'))

    # --- Text фильтры (поиск по подстроке) ---
    cipher_search = params.get('cipher_search', '').strip()
    if cipher_search:
        queryset = queryset.filter(cipher__icontains=cipher_search)

    object_id_search = params.get('object_id_search', '').strip()
    if object_id_search:
        queryset = queryset.filter(object_id__icontains=object_id_search)

    pi_number_search = params.get('pi_number_search', '').strip()
    if pi_number_search:
        queryset = queryset.filter(pi_number__icontains=pi_number_search)

    doc_number_search = params.get('accompanying_doc_number_search', '').strip()
    if doc_number_search:
        queryset = queryset.filter(accompanying_doc_number__icontains=doc_number_search)

    # --- Date range фильтры ---
    for date_field in ('registration_date', 'deadline', 'manufacturing_deadline'):
        date_from = params.get(f'{date_field}_from')
        date_to = params.get(f'{date_field}_to')
        if date_from:
            queryset = queryset.filter(**{f'{date_field}__gte': date_from})
        if date_to:
            queryset = queryset.filter(**{f'{date_field}__lte': date_to})

    return queryset


def _count_active_filters(params):
    """Подсчитывает количество активных фильтров."""
    count = 0
    filter_keys = [
        'status', 'workshop_status', 'laboratory', 'client', 'contract',
        'standards', 'accreditation_area', 'test_type', 'report_type',
        'further_movement', 'registered_by', 'verified_by',
        'manufacturing', 'uzk_required',
        'cipher_search', 'object_id_search', 'pi_number_search',
        'accompanying_doc_number_search',
    ]
    for key in filter_keys:
        if params.getlist(key):
            count += 1
    for date_field in ('registration_date', 'deadline', 'manufacturing_deadline'):
        if params.get(f'{date_field}_from') or params.get(f'{date_field}_to'):
            count += 1
    return count


def _get_filter_options_for_queryset(queryset):
    """
    Возвращает доступные варианты фильтров для данного queryset.
    Используется для каскадных фильтров.
    """
    from core.models import SampleStatus, WorkshopStatus, ReportType, FurtherMovement

    base_qs = queryset.order_by().select_related(None).prefetch_related(None)

    options = {}

    # Статусы
    existing_statuses = set(base_qs.values_list('status', flat=True).distinct())
    options['status'] = [
        {'value': s.value, 'label': s.label}
        for s in SampleStatus
        if s.value in existing_statuses
    ]

    # Workshop status
    existing_ws = set(base_qs.exclude(workshop_status__isnull=True).values_list('workshop_status', flat=True).distinct())
    options['workshop_status'] = [
        {'value': s.value, 'label': s.label}
        for s in WorkshopStatus
        if s.value in existing_ws
    ]

    # Лаборатории
    labs = base_qs.values_list(
        'laboratory_id', 'laboratory__code_display', 'laboratory__name'
    ).distinct().order_by('laboratory__code_display')
    options['laboratory'] = [
        {'value': str(l[0]), 'label': f"{l[1]} — {l[2]}"}
        for l in labs if l[0]
    ]

    # Заказчики
    clients = base_qs.values_list(
        'client_id', 'client__name'
    ).distinct().order_by('client__name')
    options['client'] = [
        {'value': str(c[0]), 'label': c[1]}
        for c in clients if c[0]
    ]

    # Договоры
    contracts = base_qs.exclude(contract__isnull=True).values_list(
        'contract_id', 'contract__number'
    ).distinct().order_by('contract__number')
    options['contract'] = [
        {'value': str(c[0]), 'label': c[1]}
        for c in contracts if c[0]
    ]

    # Стандарты
    stds = base_qs.values_list(
        'standards__id', 'standards__code'
    ).distinct().order_by('standards__code')
    options['standards'] = [
        {'value': str(s[0]), 'label': s[1]}
        for s in stds if s[0]
    ]

    # Области аккредитации
    areas = base_qs.values_list(
        'accreditation_area_id', 'accreditation_area__code'
    ).distinct().order_by('accreditation_area__code')
    options['accreditation_area'] = [
        {'value': str(a[0]), 'label': a[1]}
        for a in areas if a[0]
    ]

    # Типы испытаний
    test_types = base_qs.exclude(
        test_type=''
    ).values_list('test_type', flat=True).distinct().order_by('test_type')
    options['test_type'] = [
        {'value': t, 'label': t}
        for t in test_types
    ]

    # Report type
    existing_rt = set(base_qs.values_list('report_type', flat=True).distinct())
    options['report_type'] = [
        {'value': r.value, 'label': r.label}
        for r in ReportType
        if r.value in existing_rt
    ]

    # Further movement
    existing_fm = set(base_qs.exclude(further_movement='').values_list('further_movement', flat=True).distinct())
    options['further_movement'] = [
        {'value': f.value, 'label': f.label}
        for f in FurtherMovement
        if f.value and f.value in existing_fm
    ]

    # Registered by
    reg_by = base_qs.values_list(
        'registered_by_id', 'registered_by__last_name', 'registered_by__first_name'
    ).distinct().order_by('registered_by__last_name', 'registered_by__first_name')
    options['registered_by'] = [
        {'value': str(r[0]), 'label': f"{r[2]} {r[1]}".strip()}
        for r in reg_by if r[0]
    ]

    # Verified by (nullable)
    has_unverified = base_qs.filter(verified_by__isnull=True).exists()
    ver_by = base_qs.exclude(verified_by__isnull=True).values_list(
        'verified_by_id', 'verified_by__last_name', 'verified_by__first_name'
    ).distinct().order_by('verified_by__last_name', 'verified_by__first_name')
    verified_options = []
    if has_unverified:
        verified_options.append({'value': '__none__', 'label': '⏳ Ожидает проверки'})
    for v in ver_by:
        verified_options.append({'value': str(v[0]), 'label': f"{v[2]} {v[1]}".strip()})
    options['verified_by'] = verified_options

    return options


def _apply_sorting(samples, sort_field, sort_dir, user_role):
    """Применяет сортировку к queryset."""
    if sort_field and sort_field in DISPLAYABLE_COLUMNS_DICT:
        sort_map = {
            'client': 'client__name',
            'contract': 'contract__number',
            'standards': 'standard__code',
            'laboratory': 'laboratory__code_display',
            'accreditation_area': 'accreditation_area__code',
            'registered_by': 'registered_by__last_name',
            'verified_by': 'verified_by__last_name',
            # ⭐ v3.84.0: report_preparers — M2M. Сортируем по фамилии первого
            # (по id связи) preparer'а. Не идеально для кейсов с несколькими
            # preparer'ами, но для 99% записей с одним человеком — норм.
            'report_preparers': 'report_preparers__last_name',
            'protocol_checked_by': 'protocol_checked_by__last_name',
        }
        db_field = sort_map.get(sort_field, sort_field)
        if sort_dir == 'desc':
            db_field = f'-{db_field}'
        return samples.order_by(db_field)
    else:
        if user_role in ('CLIENT_MANAGER', 'CLIENT_DEPT_HEAD', 'SYSADMIN'):
            return samples.order_by('-registration_date', '-sequence_number')
        return samples.order_by('-sequence_number')


def _get_export_value(sample, column_code):
    """Возвращает значение поля образца для экспорта в XLSX."""
    if column_code == 'sequence_number':
        return sample.sequence_number
    elif column_code == 'registration_date':
        return sample.registration_date
    elif column_code == 'laboratory':
        return sample.laboratory.code_display if sample.laboratory else ''
    elif column_code == 'cipher':
        return sample.cipher
    elif column_code == 'client':
        return sample.client.name if sample.client else ''
    elif column_code == 'contract':
        return sample.contract.number if sample.contract else ''
    elif column_code == 'contract_date':
        return sample.contract_date
    elif column_code == 'standards':
        stds = sample.standards.all()
        return ', '.join(s.code for s in stds) if stds else ''
    elif column_code == 'test_type':
        return sample.test_type or ''
    elif column_code == 'test_code':
        return sample.test_code or ''
    elif column_code == 'accreditation_area':
        return sample.accreditation_area.code if sample.accreditation_area else ''
    elif column_code == 'deadline':
        return sample.deadline
    elif column_code == 'manufacturing_deadline':
        return sample.manufacturing_deadline
    elif column_code == 'registered_by':
        return sample.registered_by.short_name if sample.registered_by else ''
    elif column_code == 'verified_by':
        return sample.verified_by.short_name if sample.verified_by else 'Ожидает'
    elif column_code == 'pi_number':
        return sample.pi_number or ''
    elif column_code == 'protocol_checked_by':
        return sample.protocol_checked_by.short_name if sample.protocol_checked_by else ''
    elif column_code == 'operators':
        ops = sample.operators.all()
        return ', '.join(op.short_name for op in ops) if ops else ''
    elif column_code == 'status':
        return sample.get_status_display()
    elif column_code == 'workshop_status':
        return sample.get_workshop_status_display() if sample.workshop_status else ''
    elif column_code == 'manufacturing':
        return 'Да' if sample.manufacturing else 'Нет'
    elif column_code == 'uzk_required':
        return 'Да' if sample.uzk_required else 'Нет'
    elif column_code == 'replacement_protocol_required':
        return 'Да' if sample.replacement_protocol_required else 'Нет'
    elif column_code == 'report_type':
        # ⭐ v3.32.0: report_type — запятая-разделённый список
        from core.models.sample import ReportType
        if sample.report_type:
            labels_map = dict(ReportType.choices)
            return ', '.join(
                labels_map.get(rt, rt) for rt in sample.report_type.split(',')
            )
        return ''
    elif column_code == 'further_movement':
        return sample.get_further_movement_display() or ''
    elif column_code == 'sample_count':
        return sample.sample_count_display
    elif column_code == 'additional_sample_count':
        return sample.additional_sample_count or 0
    elif column_code == 'working_days':
        return sample.working_days
    elif column_code == 'report_preparers':
        # ⭐ v3.84.0: M2M — перечисляем всех через запятую.
        # .all() использует prefetch_related (см. select в journal_samples view).
        preparers = sample.report_preparers.all()
        return ', '.join(u.short_name for u in preparers) if preparers else ''
    # DateTime поля
    elif column_code in ('conditioning_start_datetime', 'conditioning_end_datetime',
                         'testing_start_datetime', 'testing_end_datetime',
                         'report_prepared_date', 'manufacturing_completion_date'):
        val = getattr(sample, column_code, None)
        if val:
            return localtime(val).strftime('%d.%m.%Y %H:%M')
        return ''
    # Date поля
    elif column_code in ('protocol_issued_date', 'protocol_printed_date',
                         'replacement_protocol_issued_date', 'sample_received_date'):
        val = getattr(sample, column_code, None)
        return val if val else ''
    # Текстовые поля
    else:
        val = getattr(sample, column_code, None)
        return val if val else ''


# ─────────────────────────────────────────────────────────────
# Views
# ─────────────────────────────────────────────────────────────

@login_required
def journal_samples(request):
    """Журнал образцов: серверная пагинация, фильтрация, кастомизация столбцов."""

    if not PermissionChecker.has_journal_access(request.user, 'SAMPLES'):
        messages.error(request, 'У вас нет доступа к журналу образцов')
        return redirect('workspace_home')

    user = request.user
    user_role = user.role

    # ⭐ v3.81.1: явный сброс через ?_reset=1
    # Используется resetFilters и applyFilters (когда набор фильтров пуст).
    # Атомарная серверная очистка saved исключает race condition с фоновым
    # IIFE-автосохранением на покидаемой странице. Без этого reset иногда
    # «не сбрасывает», а apply с пустой формой уводит редиректом на saved.
    if request.GET.get('_reset') == '1':
        prefs = user.ui_preferences or {}
        if (prefs.get('journal_filters') or {}).get('SAMPLES'):
            prefs.setdefault('journal_filters', {})['SAMPLES'] = ''
            user.ui_preferences = prefs
            user.save(update_fields=['ui_preferences'])
        clean = request.GET.copy()
        clean.pop('_reset', None)
        target = request.path
        if clean:
            target = f"{target}?{clean.urlencode()}"
        return redirect(target)

    # ⭐ v3.81.0: Подтягиваем сохранённые фильтры, если URL без GET-параметров.
    # Пустой request.GET = "зашёл с нуля" (клик в меню, прямой URL).
    # URL с параметрами (в т.ч. после reset) — применяем как есть, без редиректа.
    if not request.GET:
        prefs = user.ui_preferences or {}
        saved_qs = (prefs.get('journal_filters') or {}).get('SAMPLES', '').strip()
        if saved_qs:
            return redirect(f"{request.path}?{saved_qs}")

    # ─── Queryset ───
    samples = _build_base_queryset(user)

    samples = samples.select_related(
        # ⭐ v3.84.0: report_prepared_by (FK) удалён, теперь M2M report_preparers
        # (ниже в prefetch_related).
        'laboratory', 'accreditation_area', 'client', 'contract',
        'registered_by', 'verified_by', 'protocol_checked_by',
    ).prefetch_related(
        'operators', 'standards',
        'report_preparers',  # ⭐ v3.84.0
    ).distinct()

    # ─── Фильтры ───
    samples = _apply_filters(samples, request.GET, user)
    active_filter_count = _count_active_filters(request.GET)

    # ⭐ v3.34.0: По умолчанию скрываем завершённые образцы (чтобы не мешали работе)
    # Если пользователь явно выбрал фильтр по статусу — показываем всё что выбрал
    if not request.GET.getlist('status'):
        hide_statuses = ['COMPLETED']
        if user_role == 'TESTER':
            hide_statuses.append('PROTOCOL_ISSUED')
        samples = samples.exclude(status__in=hide_statuses)

    total_count = samples.count()

    # ─── Статистика ───
    stats = {'total': total_count}
    if user_role in ('CLIENT_MANAGER', 'CLIENT_DEPT_HEAD', 'SYSADMIN', 'LAB_HEAD'):
        stats['pending'] = samples.filter(status='PENDING_VERIFICATION').count()
        stats['registered'] = samples.filter(status='REGISTERED').count()
    elif user_role in ('QMS_HEAD', 'QMS_ADMIN'):
        stats['pending_protocol'] = samples.filter(status='DRAFT_READY').count()
        stats['pending_results'] = samples.filter(status='RESULTS_UPLOADED').count()
    elif user_role in WORKSHOP_ROLES:
        stats['pending'] = samples.filter(workshop_status=WorkshopStatus.IN_WORKSHOP).count()
        stats['completed'] = samples.filter(workshop_status=WorkshopStatus.COMPLETED).count()

    # ─── Сортировка ───
    sort_field = request.GET.get('sort', '')
    sort_dir = request.GET.get('dir', 'desc')
    samples = _apply_sorting(samples, sort_field, sort_dir, user_role)

    # ─── Пагинация ───
    page_number = request.GET.get('page', 1)
    try:
        per_page = int(request.GET.get('per_page', ITEMS_PER_PAGE))
        per_page = min(max(per_page, 10), 1000)  # 10..1000
    except (ValueError, TypeError):
        per_page = ITEMS_PER_PAGE
    paginator = Paginator(samples, per_page)
    page_obj = paginator.get_page(page_number)

    # ─── Столбцы ───
    available_columns = _get_user_visible_columns(user)
    selected_columns = _get_user_selected_columns(user)

    visible_columns = [
        {'code': code, 'name': DISPLAYABLE_COLUMNS_DICT.get(code, code)}
        for code in selected_columns
        if code in DISPLAYABLE_COLUMNS_DICT
    ]

    # ⭐ v3.34.0: Порядок столбцов в модалке соответствует сохранённому порядку
    selected_set = set(selected_columns)
    ordered_available = []
    # Сначала выбранные в порядке сохранения
    for code in selected_columns:
        if code in DISPLAYABLE_COLUMNS_DICT:
            ordered_available.append({
                'code': code,
                'name': DISPLAYABLE_COLUMNS_DICT[code],
                'selected': True,
            })
    # Потом невыбранные в порядке JOURNAL_DISPLAYABLE_COLUMNS
    for code in available_columns:
        if code not in selected_set and code in DISPLAYABLE_COLUMNS_DICT:
            ordered_available.append({
                'code': code,
                'name': DISPLAYABLE_COLUMNS_DICT[code],
                'selected': False,
            })
    all_available_columns = ordered_available

    # ⭐ v3.34.0: Сохранённые ширины столбцов
    prefs = user.ui_preferences or {}
    saved_column_widths = prefs.get('journal_column_widths', {}).get('SAMPLES', {})

    # ─── Фильтры ───
    filter_options = _get_filter_options_for_queryset(samples)

    available_filters = {}
    for col_code, filter_config in FILTERABLE_COLUMNS.items():
        if col_code in available_columns:
            available_filters[col_code] = {
                **filter_config,
                'options': filter_options.get(col_code, []),
            }

    current_filters = {}
    for key in FILTERABLE_COLUMNS:
        values = request.GET.getlist(key)
        if values:
            current_filters[key] = values
    for suffix in ('_search', '_from', '_to'):
        for key in request.GET:
            if key.endswith(suffix):
                current_filters[key] = request.GET.get(key)

    # ─── URL params ───
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    # ⭐ v3.32.0: Данные для вкладки «Этикетки»
    can_labels = PermissionChecker.can_view(user, 'LABELS', 'access')
    labels_samples = []
    labels_laboratories = []
    lab_label_filter = request.GET.get('labels_lab', '')
    labels_cipher_search = request.GET.get('labels_cipher', '').strip()
    if can_labels:
        labels_laboratories = list(
            Laboratory.objects.filter(is_active=True, department_type='LAB').order_by('code')
        )

        # ⭐ Сортировка этикеток
        labels_sort = request.GET.get('labels_sort', '-sequence_number')
        ALLOWED_LABELS_SORTS = {
            'sequence_number', '-sequence_number',
            'cipher', '-cipher',
            'laboratory__code', '-laboratory__code',
            'material', '-material',
            'determined_parameters', '-determined_parameters',
            'sample_count', '-sample_count',
            'label_printed', '-label_printed',
        }
        if labels_sort not in ALLOWED_LABELS_SORTS:
            labels_sort = '-sequence_number'

        labels_qs = Sample.objects.select_related(
            'laboratory', 'client', 'cutting_standard'
        ).prefetch_related('standards').exclude(
            status__in=('CANCELLED', 'PENDING_VERIFICATION', 'DRAFT')  # ⭐ v3.89.0
        ).order_by(labels_sort)

        if lab_label_filter:
            labels_qs = labels_qs.filter(laboratory__code=lab_label_filter)
        if labels_cipher_search:
            labels_qs = labels_qs.filter(cipher__icontains=labels_cipher_search)
        labels_samples = labels_qs[:200]

    # ⭐ v3.89.0: Контекст для таба «Черновики».
    can_drafts = user_role in DRAFTS_VISIBLE_ROLES
    drafts_samples = []
    drafts_owner = request.GET.get('drafts_owner', 'mine')  # mine | all
    if can_drafts:
        drafts_qs = Sample.objects.select_related(
            'laboratory', 'client', 'registered_by', 'acceptance_act',
        ).prefetch_related('standards').filter(status='DRAFT')

        if drafts_owner == 'mine':
            drafts_qs = drafts_qs.filter(registered_by=user)
        # Сортировка: старые сверху — это будущий порядок номеров при выпуске.
        drafts_qs = drafts_qs.order_by('created_at')
        drafts_samples = list(drafts_qs[:500])
        print(
            f'[DEBUG] can_drafts={can_drafts}, drafts_owner={drafts_owner}, drafts count={len(drafts_samples)}, user.id={user.id}, user.role={user.role}')  # ← TEMP

    return render(request, 'core/journal_samples.html', {
        'page_obj': page_obj,
        'samples': page_obj.object_list,
        'visible_columns': visible_columns,
        'all_available_columns': all_available_columns,
        'available_filters': json.dumps(available_filters, ensure_ascii=False),
        'current_filters': json.dumps(current_filters, ensure_ascii=False),
        'active_filter_count': active_filter_count,
        'stats': stats,
        'user': request.user,
        'journal_name': 'Журнал образцов',
        'is_workshop_view': user_role in WORKSHOP_ROLES,
        'query_string': query_string,
        'current_sort': sort_field,
        'current_dir': sort_dir,
        'total_count': total_count,
        'per_page': per_page,
        'saved_column_widths': json.dumps(saved_column_widths),
        # ⭐ v3.32.0: Этикетки
        'can_labels': can_labels,
        'labels_samples': labels_samples,
        'labels_laboratories': labels_laboratories,
        'labels_lab_filter': request.GET.get('labels_lab', ''),
        'labels_cipher_search': labels_cipher_search,
        'labels_sort': labels_sort if can_labels else '-sequence_number',
        # ⭐ v3.89.0: Черновики
        'can_drafts': can_drafts,
        'drafts_samples': drafts_samples,
        'drafts_owner': drafts_owner,
    })


@login_required
def export_journal_xlsx(request):
    """
    Экспорт журнала образцов в XLSX.
    GET-параметры: те же фильтры что и journal_samples.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not PermissionChecker.has_journal_access(request.user, 'SAMPLES'):
        return HttpResponse('Нет доступа', status=403)

    user = request.user

    # ─── Queryset ───
    samples = _build_base_queryset(user)

    samples = samples.select_related(
        # ⭐ v3.84.0: report_prepared_by (FK) удалён, теперь M2M report_preparers
        'laboratory', 'accreditation_area', 'client', 'contract',
        'registered_by', 'verified_by', 'protocol_checked_by',
    ).prefetch_related(
        'operators', 'standards',
        'report_preparers',  # ⭐ v3.84.0
    ).distinct()

    samples = _apply_filters(samples, request.GET, user)

    sort_field = request.GET.get('sort', '')
    sort_dir = request.GET.get('dir', 'desc')
    samples = _apply_sorting(samples, sort_field, sort_dir, user.role)

    # ─── Столбцы пользователя ───
    selected_columns = _get_user_selected_columns(user)
    columns = [
        (code, DISPLAYABLE_COLUMNS_DICT.get(code, code))
        for code in selected_columns
        if code in DISPLAYABLE_COLUMNS_DICT
    ]

    # ─── Генерация XLSX ───
    wb = Workbook()
    ws = wb.active
    ws.title = 'Журнал образцов'

    # Стили
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4A90E2')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    date_alignment = Alignment(horizontal='center', vertical='top')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill('solid', fgColor='F8F9FA')

    # Заголовки
    for col_idx, (code, name) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.freeze_panes = 'A2'

    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f'A1:{last_col}1'

    # Данные
    row_idx = 2
    for sample in samples:
        for col_idx, (code, name) in enumerate(columns, 1):
            value = _get_export_value(sample, code)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border

            if isinstance(value, date) and not isinstance(value, datetime):
                cell.number_format = 'DD.MM.YYYY'
                cell.alignment = date_alignment
            else:
                cell.alignment = cell_alignment

        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

        row_idx += 1

    # Автоширина
    for col_idx, (code, name) in enumerate(columns, 1):
        max_len = len(name)
        for row in range(2, min(row_idx, 52)):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    # ─── HTTP Response ───
    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'journal_samples_{now_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def journal_filter_options(request):
    """
    AJAX endpoint: возвращает доступные варианты фильтров
    с учётом уже выбранных фильтров (каскадность).
    """
    if not PermissionChecker.has_journal_access(request.user, 'SAMPLES'):
        return JsonResponse({'error': 'Нет доступа'}, status=403)

    user = request.user

    samples = _build_base_queryset(user)

    samples = samples.select_related(
        'laboratory', 'accreditation_area', 'client', 'contract',
        'registered_by', 'verified_by',
    ).distinct()

    samples = _apply_filters(samples, request.GET, user)

    options = _get_filter_options_for_queryset(samples)
    return JsonResponse(options)


@login_required
@require_POST
def save_column_preferences(request):
    """
    AJAX endpoint: сохраняет выбранные столбцы в user.ui_preferences.
    POST body (JSON): {"columns": ["cipher", "client", "status", ...]}
    """
    try:
        data = json.loads(request.body)
        columns = data.get('columns', [])

        if not columns:
            return JsonResponse({'error': 'Список столбцов не может быть пустым'}, status=400)

        if columns == ['__reset__']:
            user = request.user
            prefs = user.ui_preferences or {}
            if 'journal_columns' in prefs and 'SAMPLES' in prefs['journal_columns']:
                del prefs['journal_columns']['SAMPLES']
                user.ui_preferences = prefs
                user.save(update_fields=['ui_preferences'])
            return JsonResponse({'status': 'ok', 'reset': True})

        available = set(_get_user_visible_columns(request.user))
        valid_columns = [c for c in columns if c in available]

        if not valid_columns:
            return JsonResponse({'error': 'Ни один из столбцов не доступен'}, status=400)

        user = request.user
        prefs = user.ui_preferences or {}
        if 'journal_columns' not in prefs:
            prefs['journal_columns'] = {}
        prefs['journal_columns']['SAMPLES'] = valid_columns
        user.ui_preferences = prefs
        user.save(update_fields=['ui_preferences'])

        return JsonResponse({'status': 'ok', 'columns': valid_columns})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ⭐ v3.34.0: Сохранение ширин столбцов журнала образцов
@login_required
@require_POST
def save_sample_column_widths(request):
    """
    AJAX endpoint: сохраняет ширины столбцов в user.ui_preferences.
    POST body (JSON): {"widths": {"cipher": 150, "client": 200, ...}}
    """
    try:
        data = json.loads(request.body)
        widths = data.get('widths', {})

        if not isinstance(widths, dict):
            return JsonResponse({'error': 'widths должен быть объектом'}, status=400)

        user = request.user
        prefs = user.ui_preferences or {}
        if 'journal_column_widths' not in prefs:
            prefs['journal_column_widths'] = {}
        prefs['journal_column_widths']['SAMPLES'] = widths
        user.ui_preferences = prefs
        user.save(update_fields=['ui_preferences'])

        return JsonResponse({'status': 'ok'})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ⭐ v3.81.0: Сохранение последних применённых фильтров журнала образцов
@login_required
@require_POST
def save_filter_preferences(request):
    """
    AJAX endpoint: сохраняет текущий query string фильтров в user.ui_preferences.
    POST body (JSON): {"filters": "status=ACTIVE&laboratory=1&laboratory=2"}
    Пустая строка = очистить сохранённые.

    Хранится как строка (не dict), чтобы при редиректе подставлялась без конвертации.
    Значения не валидируются — _apply_filters применяет whitelist ключей и ORM,
    всё лишнее молча игнорируется.
    """
    try:
        data = json.loads(request.body)
        qs = data.get('filters', '')

        if not isinstance(qs, str):
            return JsonResponse({'error': 'filters должен быть строкой'}, status=400)

        # Защита от раздутия JSONB-поля (обычный набор фильтров < 500 символов)
        if len(qs) > 4000:
            return JsonResponse({'error': 'Слишком длинная строка фильтров'}, status=400)

        user = request.user
        prefs = user.ui_preferences or {}
        if 'journal_filters' not in prefs:
            prefs['journal_filters'] = {}
        prefs['journal_filters']['SAMPLES'] = qs
        user.ui_preferences = prefs
        user.save(update_fields=['ui_preferences'])

        return JsonResponse({'status': 'ok'})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Неверный формат JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ═══════════════════════════════════════════════════════════════
# ⭐ v3.89.0: ЧЕРНОВИКИ — ВЫПУСК И УДАЛЕНИЕ
# ═══════════════════════════════════════════════════════════════

@login_required
@require_POST
def release_drafts(request):
    """
    Выпуск пула черновиков: вызывает finalize_drafts(...) с порядком
    из формы и редиректит на journal_samples с сообщением.

    POST:
        draft_ids[] — список ID черновиков в желаемом порядке
                      (после drag-and-drop в модалке).
        registration_date — дата регистрации (YYYY-MM-DD), опционально.
                            Дефолт — сегодня.
    """
    if request.user.role not in DRAFTS_VISIBLE_ROLES:
        messages.error(request, 'У вас нет прав на выпуск черновиков')
        return redirect('journal_samples')

    draft_ids_raw = request.POST.getlist('draft_ids')
    try:
        draft_ids = [int(x) for x in draft_ids_raw if x]
    except (ValueError, TypeError):
        messages.error(request, 'Некорректный список черновиков')
        return redirect('journal_samples')

    if not draft_ids:
        messages.warning(request, 'Не выбрано ни одного черновика')
        return redirect('journal_samples')

    # Дата регистрации
    reg_date_str = (request.POST.get('registration_date') or '').strip()
    registration_date = None
    if reg_date_str:
        try:
            registration_date = datetime.strptime(
                reg_date_str, '%Y-%m-%d'
            ).date()
        except ValueError:
            messages.error(request, 'Неверный формат даты регистрации')
            return redirect('journal_samples')

    # Выпуск пула
    from core.services.sample_finalization import finalize_drafts
    try:
        finalized = finalize_drafts(
            draft_ids,
            released_by=request.user,
            registration_date=registration_date,
        )
    except ValueError as e:
        messages.error(request, f'Ошибка выпуска: {e}')
        return redirect(f'{reverse("journal_samples")}#tab-drafts')
    except Exception:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('Ошибка при выпуске черновиков')
        messages.error(request, 'Не удалось выпустить пул. См. логи.')
        return redirect(f'{reverse("journal_samples")}#tab-drafts')

    # Сообщение об успехе с диапазоном номеров
    seq_numbers = sorted(s.sequence_number for s in finalized)
    if len(seq_numbers) == 1:
        range_str = f'№ {seq_numbers[0]}'
    else:
        range_str = f'№ {seq_numbers[0]}–{seq_numbers[-1]}'
    messages.success(
        request,
        f'Выпущено образцов: {len(finalized)} ({range_str}). '
        f'Дальше — обычная проверка регистрации.'
    )
    return redirect('journal_samples')


@login_required
@require_POST
def delete_draft(request, draft_id):
    """
    Физическое удаление черновика. Доступно только для DRAFT —
    обычные образцы через этот endpoint удалить нельзя.
    """
    if request.user.role not in DRAFTS_VISIBLE_ROLES:
        messages.error(request, 'У вас нет прав на удаление черновиков')
        return redirect('journal_samples')

    sample = get_object_or_404(Sample, id=draft_id)

    if sample.status != 'DRAFT':
        messages.error(
            request,
            f'Образец #{draft_id} не является черновиком (status={sample.status}). '
            f'Удалить через этот endpoint нельзя.'
        )
        return redirect('journal_samples')

    # Записываем в аудит ПЕРЕД удалением, иначе entity_id повиснет в воздухе.
    from core.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        entity_type='sample',
        entity_id=sample.id,
        action='draft_deleted',
        field_name='status',
        old_value='DRAFT',
        new_value=None,
        extra_data={
            'created_at': str(sample.created_at),
            'registered_by_id': sample.registered_by_id,
            'laboratory_id': sample.laboratory_id,
            'client_id': sample.client_id,
            'object_id': sample.object_id or None,
        },
    )

    sample.delete()
    messages.success(request, f'Черновик #{draft_id} удалён')
    return redirect(f'{reverse("journal_samples")}#tab-drafts')