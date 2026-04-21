"""
climate_views.py — Журнал контроля параметров микроклимата
v3.35.0

Маршруты:
    path('workspace/climate/', climate_views.climate_log_view, name='climate_log')
    path('workspace/climate/add/', climate_views.climate_log_add, name='climate_log_add')
    path('workspace/climate/<int:log_id>/edit/', climate_views.climate_log_edit, name='climate_log_edit')
    path('workspace/climate/<int:log_id>/delete/', climate_views.climate_log_delete, name='climate_log_delete')
"""

from datetime import date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from core.models import ClimateLog, Equipment, User
from core.models.equipment import Room
from core.services.pressure_calculator import calculate_pressure_corrected
from django.db import models

# Роли, которым разрешено вручную менять скорректированное давление
PRESSURE_EDIT_ROLES = {'SYSADMIN', 'ADMIN', 'HEAD_OF_LAB'}

# Роли, которым разрешено редактировать и удалять записи журнала климата
MANAGER_ROLES = (
    'CLIENT_MANAGER', 'CLIENT_DEPT_HEAD', 'LAB_HEAD',
    'SYSADMIN', 'QMS_HEAD', 'QMS_ADMIN', 'CTO', 'CEO',
    'WORKSHOP_HEAD',
)

ITEMS_PER_PAGE = 50


@login_required
def climate_log_view(request):
    """Журнал климата — список записей с фильтрами."""

    # Фильтры
    room_id = request.GET.get('room', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    qs = ClimateLog.objects.select_related(
        'room', 'temp_humidity_equipment', 'pressure_equipment', 'responsible'
    ).order_by('-date', '-time')

    if room_id:
        qs = qs.filter(room_id=room_id)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    # Подсчёт фильтров
    active_filters = 0
    if room_id:
        active_filters += 1
    if date_from or date_to:
        active_filters += 1

    total_count = qs.count()

    # ⭐ v3.35.0: Сводка за выбранный период (помещение + хотя бы одна дата)
    summary = None
    if room_id and (date_from or date_to):
        from django.db.models import Min, Max
        agg = qs.aggregate(
            temp_min=Min('temperature'),
            temp_max=Max('temperature'),
            hum_min=Min('humidity'),
            hum_max=Max('humidity'),
            pres_min=Min('atmospheric_pressure'),
            pres_max=Max('atmospheric_pressure'),
        )
        # Показываем только если есть хоть одно значение
        if any(v is not None for v in agg.values()):
            summary = agg

    # Пагинация
    paginator = Paginator(qs, ITEMS_PER_PAGE)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # Справочники
    rooms = Room.objects.filter(is_active=True).order_by('number')

    # СИ для климата — по boolean-флагам
    equipment_temp_humidity = Equipment.objects.filter(
        is_temp_humidity=True
    ).select_related('laboratory').order_by('accounting_number')

    equipment_pressure = Equipment.objects.filter(
        is_pressure=True
    ).select_related('laboratory').order_by('accounting_number')

    users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')

    context = {
        'page_obj': page_obj,
        'logs': page_obj.object_list,
        'total_count': total_count,
        'rooms': rooms,
        'equipment_temp_humidity': equipment_temp_humidity,
        'equipment_pressure': equipment_pressure,
        'users': users,
        'today': date.today().isoformat(),
        # Фильтры
        'f_room': room_id,
        'f_date_from': date_from,
        'f_date_to': date_to,
        'active_filters': active_filters,
        'summary': summary,
        # ⭐ v3.61.0: Разрешение на ручное редактирование давления
        'can_edit_pressure': request.user.role in PRESSURE_EDIT_ROLES,
        # Разрешение на редактирование и удаление записей
        'can_manage': request.user.role in MANAGER_ROLES,
    }
    return render(request, 'core/climate_log.html', context)


@login_required
@require_POST
def climate_log_add(request):
    """Добавить запись в журнал климата."""
    log_date = request.POST.get('date', '').strip()
    log_time = request.POST.get('time', '').strip()
    room_id = request.POST.get('room', '').strip()

    if not log_date or not log_time or not room_id:
        messages.error(request, 'Дата, время и помещение обязательны')
        return redirect('climate_log')

    temperature = request.POST.get('temperature', '').strip() or None
    humidity = request.POST.get('humidity', '').strip() or None
    temp_eq_id = request.POST.get('temp_humidity_equipment', '').strip() or None
    pressure_raw = request.POST.get('atmospheric_pressure', '').strip() or None
    pressure_eq_id = request.POST.get('pressure_equipment', '').strip() or None

    # ⭐ v3.61.0: Расчёт давления с поправками
    pressure_corrected = None
    atmospheric_pressure = pressure_raw  # по умолчанию = сырое
    if pressure_raw is not None:
        try:
            room = Room.objects.get(pk=int(room_id))
            height = room.height_above_zero
        except Room.DoesNotExist:
            height = None
        pressure_corrected = calculate_pressure_corrected(
            pressure_raw_kpa=pressure_raw,
            temperature_c=temperature,
            height_m=height,
            equipment_id=int(pressure_eq_id) if pressure_eq_id else None,
        )
        if pressure_corrected is not None:
            atmospheric_pressure = pressure_corrected

    try:
        ClimateLog.objects.create(
            date=log_date,
            time=log_time,
            room_id=int(room_id),
            temperature=temperature,
            humidity=humidity,
            temp_humidity_equipment_id=int(temp_eq_id) if temp_eq_id else None,
            atmospheric_pressure=atmospheric_pressure,
            pressure_raw=pressure_raw,
            pressure_corrected=pressure_corrected,
            pressure_manually_edited=False,
            pressure_equipment_id=int(pressure_eq_id) if pressure_eq_id else None,
            responsible=request.user,
        )
        messages.success(request, 'Запись добавлена в журнал климата')
    except Exception as e:
        messages.error(request, f'Ошибка: {e}')

    return redirect('climate_log')


@login_required
@require_POST
def climate_log_edit(request, log_id):
    """Редактировать запись журнала климата."""
    if request.user.role not in MANAGER_ROLES:
        messages.error(request, 'Недостаточно прав для редактирования записи')
        return redirect('climate_log')

    log = get_object_or_404(ClimateLog, pk=log_id)

    log_date = request.POST.get('date', '').strip()
    log_time = request.POST.get('time', '').strip()
    room_id = request.POST.get('room', '').strip()

    if not log_date or not log_time or not room_id:
        messages.error(request, 'Дата, время и помещение обязательны')
        return redirect('climate_log')

    try:
        log.date = log_date
        log.time = log_time
        log.room_id = int(room_id)
        log.temperature = request.POST.get('temperature', '').strip() or None
        log.humidity = request.POST.get('humidity', '').strip() or None

        temp_eq_id = request.POST.get('temp_humidity_equipment', '').strip()
        log.temp_humidity_equipment_id = int(temp_eq_id) if temp_eq_id else None

        pressure_raw = request.POST.get('atmospheric_pressure', '').strip() or None
        pressure_eq_id = request.POST.get('pressure_equipment', '').strip()
        log.pressure_equipment_id = int(pressure_eq_id) if pressure_eq_id else None

        responsible_id = request.POST.get('responsible', '').strip()
        log.responsible_id = int(responsible_id) if responsible_id else None

        # ⭐ v3.61.0: Расчёт давления с поправками
        if pressure_raw is not None:
            log.pressure_raw = pressure_raw
            try:
                room = Room.objects.get(pk=int(room_id))
                height = room.height_above_zero
            except Room.DoesNotExist:
                height = None
            log.pressure_corrected = calculate_pressure_corrected(
                pressure_raw_kpa=pressure_raw,
                temperature_c=log.temperature,
                height_m=height,
                equipment_id=int(pressure_eq_id) if pressure_eq_id else None,
            )

            # Проверяем ручное изменение скорректированного значения
            manual_pressure = request.POST.get('pressure_corrected_manual', '').strip()
            auto_pressure = log.pressure_corrected if log.pressure_corrected else pressure_raw
            if manual_pressure and request.user.role in PRESSURE_EDIT_ROLES:
                # Помечаем как ручное только если значение реально отличается от авто
                try:
                    from decimal import Decimal
                    manual_val = Decimal(str(manual_pressure))
                    auto_val = Decimal(str(auto_pressure)) if auto_pressure is not None else None
                    if auto_val is not None and abs(manual_val - auto_val) < Decimal('0.001'):
                        log.atmospheric_pressure = auto_pressure
                        log.pressure_manually_edited = False
                    else:
                        log.atmospheric_pressure = manual_pressure
                        log.pressure_manually_edited = True
                except Exception:
                    log.atmospheric_pressure = manual_pressure
                    log.pressure_manually_edited = True
            else:
                log.atmospheric_pressure = auto_pressure
                log.pressure_manually_edited = False
        else:
            log.pressure_raw = None
            log.pressure_corrected = None
            log.atmospheric_pressure = None
            log.pressure_manually_edited = False

        log.save()
        messages.success(request, 'Запись обновлена')
    except Exception as e:
        messages.error(request, f'Ошибка: {e}')

    return redirect('climate_log')


@login_required
@require_POST
def climate_log_delete(request, log_id):
    """Удалить запись журнала климата."""
    if request.user.role not in MANAGER_ROLES:
        messages.error(request, 'Недостаточно прав для удаления записи')
        return redirect('climate_log')

    log = get_object_or_404(ClimateLog, pk=log_id)
    log.delete()
    messages.success(request, 'Запись удалена')
    return redirect('climate_log')


# ─────────────────────────────────────────────────────────────
# ⭐ v3.35.0: Мобильная форма (QR-код)
# ─────────────────────────────────────────────────────────────

@login_required
def climate_quick_add(request):
    """Мобильная форма добавления записи (по QR-коду)."""
    from datetime import datetime as dt

    # Предзаполнение из GET-параметров (из QR-кода)
    preset_room = request.GET.get('room', '')
    preset_eq_th = request.GET.get('eq_th', '')
    preset_eq_p = request.GET.get('eq_p', '')

    rooms = Room.objects.filter(is_active=True).order_by('number')
    equipment_temp_humidity = Equipment.objects.filter(
        is_temp_humidity=True
    ).order_by('accounting_number')
    equipment_pressure = Equipment.objects.filter(
        is_pressure=True
    ).order_by('accounting_number')

    # Текущая дата/время
    now = dt.now()

    context = {
        'rooms': rooms,
        'equipment_temp_humidity': equipment_temp_humidity,
        'equipment_pressure': equipment_pressure,
        'preset_room': preset_room,
        'preset_eq_th': preset_eq_th,
        'preset_eq_p': preset_eq_p,
        'today': now.strftime('%Y-%m-%d'),
        'now_time': now.strftime('%H:%M'),
    }
    return render(request, 'core/climate_quick_add.html', context)


@login_required
@require_POST
def climate_quick_submit(request):
    """Обработка мобильной формы."""
    log_date = request.POST.get('date', '').strip()
    log_time = request.POST.get('time', '').strip()
    room_id = request.POST.get('room', '').strip()

    if not log_date or not log_time or not room_id:
        messages.error(request, 'Дата, время и помещение обязательны')
        return redirect(request.META.get('HTTP_REFERER', '/workspace/climate/quick/'))

    temperature = request.POST.get('temperature', '').strip() or None
    humidity = request.POST.get('humidity', '').strip() or None
    temp_eq_id = request.POST.get('temp_humidity_equipment', '').strip() or None
    pressure_raw = request.POST.get('atmospheric_pressure', '').strip() or None
    pressure_eq_id = request.POST.get('pressure_equipment', '').strip() or None

    try:
        room = Room.objects.get(pk=int(room_id))

        # ⭐ v3.61.0: Расчёт давления с поправками
        pressure_corrected = None
        atmospheric_pressure = pressure_raw
        if pressure_raw is not None:
            pressure_corrected = calculate_pressure_corrected(
                pressure_raw_kpa=pressure_raw,
                temperature_c=temperature,
                height_m=room.height_above_zero,
                equipment_id=int(pressure_eq_id) if pressure_eq_id else None,
            )
            if pressure_corrected is not None:
                atmospheric_pressure = pressure_corrected

        ClimateLog.objects.create(
            date=log_date,
            time=log_time,
            room_id=int(room_id),
            temperature=temperature,
            humidity=humidity,
            temp_humidity_equipment_id=int(temp_eq_id) if temp_eq_id else None,
            atmospheric_pressure=atmospheric_pressure,
            pressure_raw=pressure_raw,
            pressure_corrected=pressure_corrected,
            pressure_manually_edited=False,
            pressure_equipment_id=int(pressure_eq_id) if pressure_eq_id else None,
            responsible=request.user,
        )
        return render(request, 'core/climate_quick_success.html', {
            'room': room,
            'temperature': temperature,
            'humidity': humidity,
            'pressure_raw': pressure_raw,
            'pressure': atmospheric_pressure,
        })
    except Exception as e:
        messages.error(request, f'Ошибка: {e}')
        return redirect(request.META.get('HTTP_REFERER', '/workspace/climate/quick/'))


@login_required
def climate_qr_codes(request):
    """Страница генерации QR-кодов (для SYSADMIN)."""
    if request.user.role != 'SYSADMIN':
        messages.error(request, 'Доступ только для администратора')
        return redirect('climate_log')

    rooms = Room.objects.filter(is_active=True).order_by('number')
    equipment_temp_humidity = Equipment.objects.filter(
        is_temp_humidity=True
    ).order_by('accounting_number')
    equipment_pressure = Equipment.objects.filter(
        is_pressure=True
    ).order_by('accounting_number')

    # Базовый URL для QR-кодов
    base_url = request.build_absolute_uri('/workspace/climate/quick/')

    context = {
        'rooms': rooms,
        'equipment_temp_humidity': equipment_temp_humidity,
        'equipment_pressure': equipment_pressure,
        'base_url': base_url,
    }
    return render(request, 'core/climate_qr_codes.html', context)

@login_required
def export_climate_xlsx(request):
    """
    Экспорт журнала климата в Excel.
    Учитывает текущие фильтры (room, date_from, date_to).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
 
    # ─── Queryset (как в climate_log_list) ───
    qs = ClimateLog.objects.select_related(
        'room', 'temp_humidity_equipment', 'pressure_equipment', 'responsible'
    ).order_by('-date', '-time')
 
    # ─── Применяем те же фильтры, что и на странице ───
    f_room = request.GET.get('room', '')
    f_date_from = request.GET.get('date_from', '')
    f_date_to = request.GET.get('date_to', '')
 
    if f_room:
        qs = qs.filter(room_id=f_room)
    if f_date_from:
        qs = qs.filter(date__gte=f_date_from)
    if f_date_to:
        qs = qs.filter(date__lte=f_date_to)
 
    # ─── Столбцы ───
    columns = [
        ('Дата',              14),
        ('Время',             10),
        ('Помещение',         25),
        ('Температура, °C',   16),
        ('Влажность, %',      14),
        ('СИ (темп./влажн.)', 28),
        ('Давление (сырое), кПа', 18),
        ('Давление (скорр.), кПа', 18),
        ('СИ (давление)',     28),
        ('Измерение провел',  22),
    ]
 
    # ─── Стили ───
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4A90E2')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    date_alignment = Alignment(horizontal='center', vertical='top')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    alt_fill = PatternFill('solid', fgColor='F8F9FA')
 
    # ─── Workbook ───
    wb = Workbook()
    ws = wb.active
    ws.title = 'Журнал климата'
 
    # ─── Заголовки ───
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
 
    ws.freeze_panes = 'A2'
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f'A1:{last_col}1'
 
    # ─── Данные ───
    row_idx = 2
    for log in qs:
        values = [
            log.date,
            log.time.strftime('%H:%M') if log.time else '',
            str(log.room) if log.room else '',
            log.temperature,
            log.humidity,
            str(log.temp_humidity_equipment.name) if log.temp_humidity_equipment else '',
            log.pressure_raw,
            log.atmospheric_pressure,
            str(log.pressure_equipment.name) if log.pressure_equipment else '',
            log.responsible.short_name if log.responsible else '',
        ]
 
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border
 
            if isinstance(value, date) and not isinstance(value, datetime):
                cell.number_format = 'DD.MM.YYYY'
                cell.alignment = date_alignment
            else:
                cell.alignment = cell_alignment
 
        # Чередование строк
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill
 
        row_idx += 1
 
    # ─── HTTP Response ───
    now_str = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    filename = f'climate_log_{now_str}.xlsx'
 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────
# ⭐ AJAX: оборудование для выбранного помещения
# ─────────────────────────────────────────────────────────────

from django.http import JsonResponse
from core.models.equipment import EquipmentRoom

@login_required
def climate_room_equipment(request):
    """
    GET /workspace/climate/room-equipment/?room_id=<id>
    Возвращает JSON со списком приборов (темп/влажн и давление),
    прикреплённых к указанному помещению.
    
    Прибор считается прикреплённым к помещению если:
      - equipment.room_id == room_id  (основное помещение)
      - ИЛИ существует запись EquipmentRoom(equipment, room)
    """
    room_id = request.GET.get('room_id', '').strip()

    if not room_id:
        return JsonResponse({'temp_humidity': [], 'pressure': []})

    try:
        room_id = int(room_id)
    except ValueError:
        return JsonResponse({'temp_humidity': [], 'pressure': []})

    # ID оборудования, привязанного к помещению через EquipmentRoom (доп. помещения)
    linked_ids = EquipmentRoom.objects.filter(
        room_id=room_id
    ).values_list('equipment_id', flat=True)

    # Итоговый queryset: основное помещение ИЛИ доп. помещение
    equipment_qs = Equipment.objects.filter(
        models.Q(room_id=room_id) | models.Q(id__in=linked_ids)
    ).order_by('accounting_number')

    def serialize(eq):
        label = eq.name
        if eq.factory_number:
            label += f' ({eq.factory_number})'
        return {'id': eq.id, 'label': label}

    temp_humidity = [
        serialize(eq) for eq in equipment_qs if eq.is_temp_humidity
    ]
    pressure = [
        serialize(eq) for eq in equipment_qs if eq.is_pressure
    ]

    return JsonResponse({'temp_humidity': temp_humidity, 'pressure': pressure})