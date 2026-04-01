"""
core/services/report_exporter.py

Экспорт отчёта об испытании в Excel.

Берёт оригинальный xlsx-шаблон, вырезает только блок нужного стандарта
(start_row — end_row) в новый чистый файл, заполняет данными.

Использование:
    from core.services.report_exporter import export_test_report_xlsx
    file_path = export_test_report_xlsx(report)
"""

import os
import re
import tempfile
import logging
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.cell_range import CellRange

logger = logging.getLogger(__name__)


def export_test_report_xlsx(report):
    """
    Экспортирует отчёт в xlsx.
    Вырезает только блок нужного стандарта из исходного файла.

    Returns:
        str: путь к временному xlsx-файлу
    """
    template = report.template
    if not template:
        raise ValueError('У отчёта нет привязанного шаблона')

    source = template.source
    if not source or not source.file_path:
        raise ValueError('Файл шаблона не найден')

    source_path = source.file_path
    if not os.path.exists(source_path):
        raise FileNotFoundError(f'Файл шаблона не найден: {source_path}')

    # Открываем оригинал
    wb_src = openpyxl.load_workbook(source_path)
    ws_src = wb_src[template.sheet_name]

    # Границы блока
    src_start = template.start_row
    src_end = template.end_row
    offset = src_start - 1  # строка N в оригинале → строка (N - offset) в новом файле

    # Создаём новый файл
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    std_code = template.standard.code if template.standard else 'Report'
    # Имя листа ≤ 31 символ, без спецсимволов
    ws_new.title = re.sub(r'[\\/*?\[\]:]', '', std_code)[:31]

    # Копируем блок
    _copy_block(ws_src, ws_new, src_start, src_end)

    # Заполняем данные
    _fill_header(ws_new, template, report, offset)
    _fill_specimens(ws_new, template, report, offset)
    _fill_sub_measurements(ws_new, template, report, offset)

    wb_src.close()

    # Сохраняем
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', prefix='report_', delete=False)
    tmp.close()
    wb_new.save(tmp.name)
    wb_new.close()

    return tmp.name


# ─────────────────────────────────────────────────────────────
# Копирование блока строк в новый лист
# ─────────────────────────────────────────────────────────────

def _copy_block(ws_src, ws_new, src_start, src_end):
    """
    Копирует строки src_start..src_end из ws_src в ws_new (строки 1..N).
    Переносит значения, формулы (со сдвигом), стили, размеры, merged cells.
    """
    offset = src_start - 1

    # Ширины столбцов
    for col_letter, dim in ws_src.column_dimensions.items():
        if dim.width:
            ws_new.column_dimensions[col_letter].width = dim.width

    for src_row in range(src_start, src_end + 1):
        dst_row = src_row - offset

        # Высота строки
        src_dim = ws_src.row_dimensions.get(src_row)
        if src_dim and src_dim.height:
            ws_new.row_dimensions[dst_row].height = src_dim.height

        # Ячейки
        for col in range(1, ws_src.max_column + 1):
            src_cell = ws_src.cell(row=src_row, column=col)
            dst_cell = ws_new.cell(row=dst_row, column=col)

            # Значение / формула
            value = src_cell.value
            if value and isinstance(value, str) and value.startswith('='):
                value = _adjust_formula_rows(value, offset)
            dst_cell.value = value

            # Стиль
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

    # Объединённые ячейки (только внутри блока)
    for merged in ws_src.merged_cells.ranges:
        if merged.min_row >= src_start and merged.max_row <= src_end:
            new_range = CellRange(
                min_col=merged.min_col,
                min_row=merged.min_row - offset,
                max_col=merged.max_col,
                max_row=merged.max_row - offset,
            )
            ws_new.merge_cells(str(new_range))


def _adjust_formula_rows(formula, offset):
    """
    Сдвигает номера строк в формуле на -offset.
    '=AVERAGE(N44:N46)' при offset=32 → '=AVERAGE(N12:N14)'
    """
    if offset == 0:
        return formula

    def replace_ref(match):
        col_part = match.group(1)
        row_num = int(match.group(2))
        new_row = row_num - offset
        return f'{col_part}{max(new_row, 1)}'

    return re.sub(r'([A-Z]+)(\d+)', replace_ref, formula)


# ─────────────────────────────────────────────────────────────
# Заполнение шапки
# ─────────────────────────────────────────────────────────────

def _fill_header(ws, template, report, offset):
    """Заполняет поля шапки (row смещён на -offset)."""
    header_data = report.header_data or {}
    header_config = template.header_config or {}

    # Фоллбэк: для каждого поля знаем, куда писать, если value_col не определён
    # Формат: field_key → смещение столбца относительно label
    FALLBACK_OFFSETS = {
        'date': 1,                      # A → B
        'operator': 1,                  # A → B
        'identification_number': 3,     # E → H
        'room': 1,                      # E → F
        'measuring_instruments': 1,     # A → B
        'test_equipment': 1,            # A → B
        'force_sensor': 3,              # A → D
        'traverse_speed': 3,            # A → D
        'specimen_count': 3,            # A → D
        'notes': 3,                     # A → D
        'conditions': 3,               # A → D
        'specimen_type': 3,             # A → D
    }

    for field_key, cfg in header_config.items():
        value = header_data.get(field_key)
        if not value:
            continue

        row = cfg.get('row')
        if not row:
            continue

        dst_row = row - offset
        if dst_row < 1:
            continue

        # Определяем столбец для записи
        value_col = cfg.get('value_col')
        if value_col:
            col_idx = column_index_from_string(value_col)
        else:
            # Фоллбэк: label_col + смещение
            label_col = cfg.get('col')
            if not label_col:
                continue
            label_col_idx = column_index_from_string(label_col)
            fallback_offset = FALLBACK_OFFSETS.get(field_key, 1)
            col_idx = label_col_idx + fallback_offset

        cell = ws.cell(row=dst_row, column=col_idx)

        # Не трогаем формулы
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            continue

        cell.value = value


# ─────────────────────────────────────────────────────────────
# Заполнение основной таблицы
# ─────────────────────────────────────────────────────────────

def _fill_specimens(ws, template, report, offset):
    """Заполняет строки данных (row смещён на -offset)."""
    specimens = (report.table_data or {}).get('specimens', [])
    column_config = template.column_config or []
    data_start_row = template.data_start_row - offset

    for i, spec in enumerate(specimens):
        row = data_start_row + i
        values = spec.get('values', {})

        for col_cfg in column_config:
            code = col_cfg['code']
            col_idx = column_index_from_string(col_cfg['col_letter'])
            col_type = col_cfg.get('type', 'INPUT')

            if code == 'specimen_number':
                ws.cell(row=row, column=col_idx).value = i + 1
                continue

            if code == 'marking':
                ws.cell(row=row, column=col_idx).value = spec.get('marking', '')
                continue

            # Вычисляемые — не трогаем формулы
            if col_type in ('CALCULATED', 'SUB_AVG'):
                existing = ws.cell(row=row, column=col_idx).value
                if existing and isinstance(existing, str) and existing.startswith('='):
                    continue
                val = values.get(code)
                if val is not None:
                    ws.cell(row=row, column=col_idx).value = val
                continue

            # INPUT / TEXT
            val = values.get(code)
            if val is not None:
                ws.cell(row=row, column=col_idx).value = val


# ─────────────────────────────────────────────────────────────
# Заполнение промежуточных замеров
# ─────────────────────────────────────────────────────────────

def _fill_sub_measurements(ws, template, report, offset):
    """Заполняет боковую таблицу замеров (row смещён на -offset)."""
    sub_config = template.sub_measurements_config
    if not sub_config:
        return

    specimens = (report.table_data or {}).get('specimens', [])
    sub_columns = sub_config.get('columns', [])
    mpp = sub_config.get('measurements_per_specimen', 3)

    sub_header_row = sub_config.get('header_row')
    if not sub_header_row:
        return

    sub_data_start = sub_header_row - offset + 1

    for i, spec in enumerate(specimens):
        sub_data = spec.get('sub_measurements', {})

        for col_cfg in sub_columns:
            code = col_cfg['code']
            col_idx = column_index_from_string(col_cfg['col_letter'])
            measurements = sub_data.get(code, [])

            for m in range(mpp):
                row = sub_data_start + (i * mpp) + m
                val = measurements[m] if m < len(measurements) else None
                if val is not None:
                    ws.cell(row=row, column=col_idx).value = val