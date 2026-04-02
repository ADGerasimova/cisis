"""
core/services/template_parser.py

Парсер xlsx-файлов с шаблонами таблиц для отчётов об испытаниях.

Логика:
1. Сканирует все листы xlsx-файла
2. Находит блоки шаблонов по маркеру "Дата:" в первых столбцах
3. Для каждого блока извлекает:
   - Код стандарта (из ячейки рядом с "НД:")
   - Строку заголовков (по "№ образца")
   - Конфигурацию столбцов (INPUT / CALCULATED / SUB_AVG / TEXT)
   - Конфигурацию шапки (какие поля и где)
   - Статистические строки (Среднее, Ст.откл, CV%, Дов.интервал)
   - Боковую таблицу промежуточных замеров
4. Создаёт записи в report_template_sources + report_template_index

Использование:
    from core.services.template_parser import parse_template_file
    result = parse_template_file(file_path, laboratory_id, uploaded_by_id)

Пробный запуск (без записи в БД):
    from core.services.template_parser import parse_template_file_dry_run
    results = parse_template_file_dry_run('/path/to/file.xlsx')
"""

import re
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------

BLOCK_START_MARKER = 'Дата:'

HEADER_FIELDS = {
    'Дата:': 'date',
    'Оператор:': 'operator',
    'Помещение': 'room',
    'Идентификационный номер:': 'identification_number',
    'СИ': 'measuring_instruments',
    'ИО': 'testing_equipment',
    'Датчик силы:': 'force_sensor',
    'Скорость траверсы:': 'traverse_speed',
    'Кол-во образцов:': 'specimen_count',
    'Количество образцов:': 'specimen_count',
    'Примечания:': 'notes',
    'Условия испытаний:': 'conditions',
    'Тип образцов': 'specimen_type',
    'Тип образца:': 'specimen_type',
}

STAT_MARKERS = {
    'Среднее арифметическое значение': 'MEAN',
    'Стандартное отклонение': 'STDEV',
    'Коэффициент вариации, %': 'CV',
    'Коэффициент вариации,%': 'CV',
    'Границы доверительного интервала': 'CONFIDENCE',
}

MAX_SCAN_COLS = 25


# ---------------------------------------------------------------------------
# Основная функция
# ---------------------------------------------------------------------------

def parse_template_file(file_path, laboratory_id=None, uploaded_by_id=None, description=None):
    """
    Парсит xlsx-файл и создаёт записи в БД.

    Returns:
        dict с ключами source_id, templates_created, templates_skipped, details
    """
    from core.models import ReportTemplateSource

    wb = load_workbook(file_path, data_only=False)
    file_name = file_path.split('/')[-1] if '/' in file_path else file_path

    source = ReportTemplateSource.objects.create(
        laboratory_id=laboratory_id,
        file_name=file_name,
        file_path=file_path,
        uploaded_by_id=uploaded_by_id,
        description=description or '',
    )

    result = {
        'source_id': source.id,
        'templates_created': 0,
        'templates_skipped': 0,
        'details': [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blocks = _find_template_blocks(ws)
        logger.info(f"Sheet '{sheet_name}': found {len(blocks)} template blocks")

        for block in blocks:
            detail = _process_block(ws, sheet_name, block, source, result)
            result['details'].append(detail)

    wb.close()
    logger.info(
        f"Parsed '{file_name}': {result['templates_created']} created, "
        f"{result['templates_skipped']} skipped"
    )
    return result


# ---------------------------------------------------------------------------
# Поиск блоков шаблонов на листе
# ---------------------------------------------------------------------------

def _find_template_blocks(ws):
    """
    Находит все блоки шаблонов на листе.
    Блок начинается с "Дата:" в первых 5 столбцах.
    """
    blocks = []

    for row in range(1, ws.max_row + 1):
        for col in range(1, 6):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if cell_value.strip() == BLOCK_START_MARKER:
                    blocks.append({'start_row': row, 'start_col': col})
                    break

    for i, block in enumerate(blocks):
        if i + 1 < len(blocks):
            block['end_row'] = blocks[i + 1]['start_row'] - 1
        else:
            block['end_row'] = ws.max_row

    return blocks


# ---------------------------------------------------------------------------
# Обработка одного блока
# ---------------------------------------------------------------------------

def _process_block(ws, sheet_name, block, source, result):
    """Обрабатывает один блок и создаёт запись в report_template_index."""
    from core.models import ReportTemplateIndex, Standard

    start_row = block['start_row']
    end_row = block['end_row']
    base_col = block['start_col']
    detail = {'start_row': start_row, 'end_row': end_row}

    # 1. Код стандарта
    standard_code = _find_standard_code(ws, start_row, end_row, base_col)
    detail['standard_code'] = standard_code

    if not standard_code:
        detail['status'] = 'error'
        detail['error'] = 'Код стандарта не найден'
        result['templates_skipped'] += 1
        return detail

    # 2. Стандарт в БД
    standard = (
        Standard.objects.filter(code=standard_code).first()
        or Standard.objects.filter(code__iexact=standard_code.strip()).first()
    )
    if not standard:
        detail['status'] = 'skipped'
        detail['error'] = f'Стандарт "{standard_code}" не найден в БД'
        result['templates_skipped'] += 1
        return detail

    # 3. Проверяем текущую версию
    current_template = (
        ReportTemplateIndex.objects
        .filter(standard=standard, is_current=True)
        .first()
    )

    # 4. Строка заголовков
    header_row = _find_header_row(ws, start_row, end_row, base_col)
    if not header_row:
        detail['status'] = 'error'
        detail['error'] = 'Строка заголовков не найдена'
        result['templates_skipped'] += 1
        return detail

    data_start_row = header_row + 1

    # 5. Конфигурации
    header_config = _extract_header_config(ws, start_row, header_row, base_col)
    column_config = _extract_column_config(ws, header_row, data_start_row, end_row, base_col)
    stats_start_row, statistics_config = _extract_statistics_config(
        ws, data_start_row, end_row, base_col
    )
    # В _process_block, после извлечения sub_measurements_config:

    sub_measurements_config = _extract_sub_measurements_config(ws, start_row, header_row, end_row)

    # Определяем layout_type
    if sub_measurements_config:
        # Проверяем, есть ли в боковой таблице вычисляемые столбцы
        has_calculated_in_sub = any(
            col.get('type') in ('CALCULATED', 'SUB_AVG') 
            for col in sub_measurements_config.get('columns', [])
        )
        layout_type = 'B_CALC' if has_calculated_in_sub else 'B'
    else:
        layout_type = 'A'

    # 6. Сравниваем с текущей версией — если column_config не изменился, пропускаем
    if current_template:
        import json
        old_cols = json.dumps(current_template.column_config, sort_keys=True, ensure_ascii=False)
        new_cols = json.dumps(column_config, sort_keys=True, ensure_ascii=False)
        if old_cols == new_cols:
            detail['status'] = 'skipped'
            detail['error'] = f'Шаблон для "{standard_code}" не изменился (v{current_template.version})'
            result['templates_skipped'] += 1
            return detail

    # 7. Определяем номер версии
    new_version = (current_template.version + 1) if current_template else 1

    # 8. Создание новой версии
    try:
        # Снимаем is_current со старой версии
        if current_template:
            ReportTemplateIndex.objects.filter(
                standard=standard, is_current=True
            ).update(is_current=False)

        template = ReportTemplateIndex.objects.create(
            standard=standard,
            source=source,
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            header_row=header_row,
            data_start_row=data_start_row,
            stats_start_row=stats_start_row,
            column_config=column_config,
            header_config=header_config,
            statistics_config=statistics_config,
            sub_measurements_config=sub_measurements_config,
            layout_type=layout_type,
            version=new_version,
            is_current=True,
            changes_description=f'Загружено из {source.file_name}' if new_version == 1
                else f'Обновлено из {source.file_name} (v{current_template.version} → v{new_version})',
        )
        detail['status'] = 'created' if new_version == 1 else 'updated'
        detail['template_id'] = template.id
        detail['version'] = new_version
        detail['columns_count'] = len(column_config)
        detail['has_statistics'] = bool(statistics_config)
        detail['has_sub_measurements'] = sub_measurements_config is not None
        result['templates_created'] += 1
    except Exception as e:
        detail['status'] = 'error'
        detail['error'] = str(e)
        result['templates_skipped'] += 1
        logger.exception(f"Error creating template for {standard_code}")

    return detail


# ---------------------------------------------------------------------------
# Извлечение кода стандарта
# ---------------------------------------------------------------------------

def _find_standard_code(ws, start_row, end_row, base_col):
    """Ищет "НД:" и берёт код из соседней ячейки справа."""
    search_end = min(start_row + 8, end_row + 1)
    for row in range(start_row, search_end):
        for col in range(base_col, base_col + 5):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                text = cell_value.strip()
                if text.startswith('НД') and ':' in text:
                    code_cell = ws.cell(row=row, column=col + 1).value
                    if code_cell:
                        return str(code_cell).strip()
                    parts = text.split(':', 1)
                    if len(parts) > 1 and parts[1].strip():
                        return parts[1].strip()
    return None


# ---------------------------------------------------------------------------
# Поиск строки заголовков таблицы
# ---------------------------------------------------------------------------

def _find_header_row(ws, start_row, end_row, base_col):
    """
    Ищет строку с заголовками основной таблицы.
    Маркер: "№ образца". Берём строку с максимальным числом заголовков
    (отличает основную таблицу от боковой таблицы замеров).
    """
    best_row = None
    best_count = 0

    for row in range(start_row, end_row + 1):
        for col in range(base_col, base_col + 3):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if '№ образца' in cell_value.strip():
                    non_empty = sum(
                        1 for c in range(col, col + MAX_SCAN_COLS)
                        if ws.cell(row=row, column=c).value is not None
                    )
                    if non_empty > best_count:
                        best_count = non_empty
                        best_row = row

    return best_row if best_count >= 4 else None


# ---------------------------------------------------------------------------
# Конфигурация шапки
# ---------------------------------------------------------------------------

def _extract_header_config(ws, start_row, header_row, base_col):
    """Извлекает поля шапки: {field_key: {row, col, label, value_col, row_offset}}."""
    config = {}
    for row in range(start_row, header_row):
        for col in range(base_col, base_col + 10):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value or not isinstance(cell_value, str):
                continue
            text = cell_value.strip()
            for marker, field_key in HEADER_FIELDS.items():
                if text == marker or text.startswith(marker):
                    value_col = None
                    for vc in range(col + 1, col + 8):
                        if ws.cell(row=row, column=vc).value is not None:
                            value_col = get_column_letter(vc)
                            break
                    config[field_key] = {
                        'row': row,
                        'col': get_column_letter(col),
                        'label': marker,
                        'value_col': value_col,
                        'row_offset': row - start_row,
                    }
                    break
    return config


# ---------------------------------------------------------------------------
# Конфигурация столбцов таблицы
# ---------------------------------------------------------------------------

def _extract_column_config(ws, header_row, data_start_row, end_row, base_col):
    """
    Извлекает конфигурацию столбцов основной таблицы.
    Каждый столбец: {code, name, unit, col_letter, type, formula, header_text}
    """
    columns = []
    for col in range(base_col, base_col + MAX_SCAN_COLS):
        header_value = ws.cell(row=header_row, column=col).value
        if header_value is None:
            continue
        header_text = str(header_value).strip()
        if not header_text:
            continue

        col_letter = get_column_letter(col)
        name, unit = _parse_header_name(header_text)
        first_data_cell = ws.cell(row=data_start_row, column=col).value
        col_type, formula = _detect_column_type(first_data_cell, header_text)
        code = _generate_column_code(name, header_text, col_letter)

        columns.append({
            'code': code,
            'name': name,
            'unit': unit,
            'col_letter': col_letter,
            'type': col_type,
            'formula': formula,
            'header_text': header_text,
        })
    return columns


def _parse_header_name(header_text):
    """Парсит "σ, МПа" → ("σ", "МПа")."""
    if ',' in header_text:
        parts = header_text.split(',', 1)
        return parts[0].strip(), parts[1].strip()
    return header_text, ''


def _detect_column_type(cell_value, header_text):
    """Определяет тип столбца: INPUT / CALCULATED / SUB_AVG / TEXT / ROW_NUMBER / SYSTEM."""
    if cell_value is None:
        text_markers = ['Маркировка', 'Характер', 'разрушения', 'Вид']
        if any(m.lower() in header_text.lower() for m in text_markers):
            return 'TEXT', None
        return 'INPUT', None

    cell_str = str(cell_value)

    if cell_str.startswith('='):
        upper = cell_str.upper()
        if 'AVERAGE' in upper and 'IFERROR' not in upper and 'STDEV' not in upper:
            return 'SUB_AVG', cell_str
        return 'CALCULATED', cell_str

    if header_text == '№ образца':
        return 'ROW_NUMBER', None
    if header_text.strip() == 'br':
        return 'SYSTEM', None

    text_markers = ['Маркировка', 'Характер', 'разрушения', 'Вид']
    if any(m.lower() in header_text.lower() for m in text_markers):
        return 'TEXT', None

    return 'INPUT', None


def _generate_column_code(name, header_text, col_letter):
    """Генерирует машинный код столбца."""
    exact_map = {
        '№ образца': 'specimen_number',
        'Маркировка образца': 'marking',
        'Характер разрушения': 'failure_mode',
        'br': 'br',
    }
    if header_text in exact_map:
        return exact_map[header_text]

    ht_lower = header_text.lower()
    if 'маркировка' in ht_lower:
        return 'marking'
    if ('характер' in ht_lower or 'вид' in ht_lower) and 'разрушен' in ht_lower:
        return 'failure_mode'

    greek_map = {
        'σ': 'sigma', 'σВ': 'sigma_v', 'σМ1': 'sigma_m1', 'σpm': 'sigma_pm',
        'Е': 'E', 'E': 'E', 'Ep': 'Ep', 'Eр': 'Ep',
        'δ': 'delta', 'ε': 'epsilon', 'εр': 'epsilon_r',
        'εобщ': 'epsilon_total',
        'ν': 'nu', 'v': 'v', 'μ12': 'mu12',
        'F': 'F_kn', 'Fmax': 'F_max', 'Pmax': 'P_max', 'P': 'P',
        'fp': 'fp', 'Ftu': 'Ftu', 'Fpm': 'Fpm',
        'hср': 'h_avg', 'bср': 'b_avg', 'dср': 'd_avg', 'aср': 'a_avg',
        'h': 'h', 'b': 'b', 'd': 'd', 'a': 'a', 'w': 'w',
        'К': 'K_coeff', 'А0': 'A0', 'S': 'S_area',
        'σ0,2%': 'sigma_02', 'D': 'D_mm', 'l': 'l_mm', 't': 't_min',
    }
    if name in greek_map:
        return greek_map[name]

    clean = re.sub(r'[^a-zA-Zа-яА-Я0-9_]', '', name)
    if clean:
        return clean.lower()
    return f'col_{col_letter}'

def _detect_column_type(cell_value, header_text):
    """Определяет тип столбца."""
    if cell_value is None:
        text_markers = ['маркировка', 'характер', 'разрушения', 'примечание']
        if any(m.lower() in header_text.lower() for m in text_markers):
            return 'TEXT', None
        return 'INPUT', None

    cell_str = str(cell_value)

    if cell_str.startswith('='):
        upper = cell_str.upper()
        
        # Проверяем на VLOOKUP/ВПР
        if 'VLOOKUP' in upper or 'ВПР' in upper:
            return 'VLOOKUP', cell_str
        
        if 'AVERAGE' in upper and 'IFERROR' not in upper:
            return 'SUB_AVG', cell_str
        return 'CALCULATED', cell_str
# ---------------------------------------------------------------------------
# Статистические строки
# ---------------------------------------------------------------------------

def _extract_statistics_config(ws, data_start_row, end_row, base_col):
    """Находит строки: Среднее, Ст.откл, CV%, Дов.интервал."""
    stats = []
    stats_start = None

    for row in range(data_start_row, end_row + 1):
        for col in range(base_col, base_col + 5):
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value or not isinstance(cell_value, str):
                continue
            text = cell_value.strip()
            for marker, stat_type in STAT_MARKERS.items():
                if marker in text:
                    if stats_start is None:
                        stats_start = row
                    formula_columns = []
                    for fc in range(base_col, base_col + MAX_SCAN_COLS):
                        fv = ws.cell(row=row, column=fc).value
                        if fv and isinstance(fv, str) and fv.startswith('='):
                            formula_columns.append({
                                'col_letter': get_column_letter(fc),
                                'formula': fv,
                            })
                    stats.append({
                        'type': stat_type,
                        'row': row,
                        'row_offset': row - data_start_row,
                        'label': text,
                        'columns': formula_columns,
                    })
                    break

    return stats_start, stats


# ---------------------------------------------------------------------------
# Промежуточные замеры (боковая таблица)
# ---------------------------------------------------------------------------

def _extract_sub_measurements_config(ws, start_row, header_row, end_row):
    """
    Ищет боковую таблицу: "№ образца", "h, мм", "b, мм" в столбцах L-V.
    
    Поддерживает вычисляемые столбцы (формулы) в боковой таблице.
    """
    for row in range(start_row, header_row):
        for col in range(12, 22):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if '№ образца' in cell_value.strip():
                    sub_header_row = row
                    sub_start_col = col

                    sub_columns = []
                    # Сканируем столбцы боковой таблицы
                    for sc in range(col + 1, col + 10):
                        sv = ws.cell(row=sub_header_row, column=sc).value
                        if sv is None:
                            break
                        sv_text = str(sv).strip()
                        if not sv_text:
                            break
                        
                        name, unit = _parse_header_name(sv_text)
                        col_letter = get_column_letter(sc)
                        
                        # Смотрим на первую строку данных, чтобы определить тип столбца
                        first_data_row = sub_header_row + 1
                        first_data_cell = ws.cell(row=first_data_row, column=sc).value
                        
                        # Определяем тип и формулу для столбца в боковой таблице
                        col_type, formula = _detect_sub_column_type(
                            first_data_cell, sv_text, name
                        )
                        
                        sub_columns.append({
                            'code': _generate_column_code(name, sv_text, col_letter),
                            'name': name,
                            'unit': unit,
                            'col_letter': col_letter,
                            'type': col_type,  # ← добавляем тип
                            'formula': formula,  # ← добавляем формулу (для CALCULATED/SUB_AVG)
                        })

                    if not sub_columns:
                        continue

                    mps = _detect_measurements_per_specimen(
                        ws, sub_header_row + 1, end_row, sub_start_col
                    )
                    return {
                        'start_col': get_column_letter(sub_start_col),
                        'header_row': sub_header_row,
                        'measurements_per_specimen': mps,
                        'columns': sub_columns,  # теперь каждый столбец имеет type и formula
                    }
    return None
# Добавьте эту функцию в template_parser.py (например, после _extract_sub_measurements_config)

def _detect_measurements_per_specimen(ws, data_row, end_row, number_col):
    """
    Определяет количество замеров на один образец в боковой таблице.
    
    Считает строки между последовательными номерами образцов (1, 2, 3...)
    """
    first_row = None
    second_row = None

    for row in range(data_row, min(data_row + 20, end_row + 1)):
        cell = ws.cell(row=row, column=number_col).value
        if cell is not None:
            # Если нашли число (номер образца)
            if isinstance(cell, (int, float)):
                if first_row is None:
                    first_row = row
                elif second_row is None and row > first_row:
                    second_row = row
                    break
            # Если нашли формулу (например, =A1+1)
            elif isinstance(cell, str) and cell.startswith('='):
                if first_row is not None and second_row is None:
                    second_row = row
                    break

    if first_row and second_row and second_row > first_row:
        return second_row - first_row
    
    # По умолчанию 3 замера
    return 3   

def _detect_sub_column_type(cell_value, header_text, name):
    """
    Определяет тип столбца в боковой таблице.
    
    Возвращает:
        ('INPUT', None) - ручной ввод
        ('TEXT', None) - текстовое поле
        ('CALCULATED', formula) - вычисляемый столбец
        ('SUB_AVG', formula) - среднее из других замеров
    """
    if cell_value is None:
        # Если нет данных, смотрим по заголовку
        text_markers = ['маркировка', 'примечание', 'комментарий']
        if any(m in header_text.lower() for m in text_markers):
            return 'TEXT', None
        
        # По умолчанию - ручной ввод
        return 'INPUT', None
    
    cell_str = str(cell_value)
    
    # Если начинается с формулы
    if cell_str.startswith('='):
        upper = cell_str.upper()
        # Среднее арифметическое
        if 'AVERAGE' in upper and 'IFERROR' not in upper and 'STDEV' not in upper:
            return 'SUB_AVG', cell_str
        # Любая другая формула
        return 'CALCULATED', cell_str
    
    # Текстовые поля
    text_markers = ['маркировка', 'примечание', 'комментарий', 'характер']
    if any(m in header_text.lower() for m in text_markers):
        return 'TEXT', None
    
    # По умолчанию - ручной ввод
    return 'INPUT', None


# ---------------------------------------------------------------------------
# Пробный запуск (без записи в БД)
# ---------------------------------------------------------------------------

def parse_template_file_dry_run(file_path):
    """
    Пробный запуск парсера БЕЗ записи в БД.

    Использование из Django shell:
        from core.services.template_parser import parse_template_file_dry_run
        results = parse_template_file_dry_run('/path/to/file.xlsx')
        for r in results:
            print(r['standard_code'], r['columns_count'], r['statistics'])
    """
    wb = load_workbook(file_path, data_only=False)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blocks = _find_template_blocks(ws)

        for block in blocks:
            start_row = block['start_row']
            end_row = block['end_row']
            base_col = block['start_col']

            standard_code = _find_standard_code(ws, start_row, end_row, base_col)
            header_row = _find_header_row(ws, start_row, end_row, base_col)

            if not header_row:
                results.append({
                    'standard_code': standard_code,
                    'error': 'Header row not found',
                })
                continue

            data_start_row = header_row + 1
            header_config = _extract_header_config(ws, start_row, header_row, base_col)
            column_config = _extract_column_config(
                ws, header_row, data_start_row, end_row, base_col
            )
            stats_start_row, statistics_config = _extract_statistics_config(
                ws, data_start_row, end_row, base_col
            )
            sub_measurements_config = _extract_sub_measurements_config(
                ws, start_row, header_row, end_row
            )

            results.append({
                'sheet': sheet_name,
                'standard_code': standard_code,
                'rows': f'{start_row}-{end_row}',
                'header_row': header_row,
                'columns': column_config,
                'columns_count': len(column_config),
                'header_fields': list(header_config.keys()),
                'statistics': [s['type'] for s in statistics_config],
                'has_sub_measurements': sub_measurements_config is not None,
                'sub_measurements': sub_measurements_config,
            })

    wb.close()
    return results